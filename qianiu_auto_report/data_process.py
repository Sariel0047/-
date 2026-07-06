"""
数据处理模块。
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
from typing import Any, Optional

import pandas as pd

from qianiu_auto_report.utils import get_latest_file, get_previous_date


class DataProcessor:
    """
    数据清洗与整理处理器。
    """

    SUPPORTED_EXCEL_SUFFIXES = (".xlsx", ".xls", ".xlsm")
    SUPPORTED_TABLE_SUFFIXES = SUPPORTED_EXCEL_SUFFIXES + (".csv",)
    DATE_COLUMN_CANDIDATES = (
        "退款完结时间",
        "退款成功时间",
        "完结时间",
        "申请时间",
        "售后申请时间",
        "创建时间",
        "申请日期",
        "日期",
        "时间",
    )
    AMOUNT_COLUMN_CANDIDATES = (
        "退款总额",
        "退款总金额",
        "退款金额",
        "金额",
        "申请退款金额",
        "实退金额",
    )
    REFUND_TYPE_COLUMN_CANDIDATES = (
        "售后类型",
        "服务类型",
        "退款类型",
        "申请类型",
        "维权类型",
    )
    GOODS_STATUS_COLUMN_CANDIDATES = (
        "货物状态",
        "收货状态",
        "退货状态",
        "物流状态",
    )
    CATEGORY_STRUCTURE = {
        "已发货仅退款": ("未收到货", "已收到货"),
        "未发货仅退款": ("未发货",),
        "退货退款": ("已寄回",),
    }
    DOUYIN_REFUND_DETAIL_SHEET_NAME = "本店数据"
    DOUYIN_AFTER_SALE_TYPE_COLUMN = "售后类型"
    DOUYIN_AFTER_SALE_AMOUNT_COLUMN = "退商品金额（元）"
    DOUYIN_AFTER_SALE_FINISHED_TIME_COLUMN = "售后完结时间"
    DOUYIN_REFUND_METRIC_COLUMNS = {
        "refund_total_order_count": "全部退款阶段-退款订单数",
        "refund_total_amount": "全部退款阶段-退款金额",
        "pre_shipment_refund_order_count": "发货前退款阶段-退款订单数",
        "pre_shipment_refund_amount": "发货前退款阶段-退款金额",
        "unreceived_refund_order_count": "未收货退款阶段-退款订单数",
        "unreceived_refund_amount": "未收货退款阶段-退款金额",
        "received_refund_order_count": "已收货退款阶段-退款订单数",
        "received_refund_amount": "已收货退款阶段-退款金额",
        "return_refund_order_count": "已收货退货退款阶段-退款订单数",
        "return_refund_amount": "已收货退货退款阶段-退款金额",
    }
    DOUYIN_ZERO_REFUND_METRICS = {
        "refund_total_order_count": 0,
        "refund_total_amount": 0.0,
        "pre_shipment_refund_order_count": 0,
        "pre_shipment_refund_amount": 0.0,
        "unreceived_refund_order_count": 0,
        "unreceived_refund_amount": 0.0,
        "received_refund_order_count": 0,
        "received_refund_amount": 0.0,
        "return_refund_order_count": 0,
        "return_refund_amount": 0.0,
    }
    TMALL_SOLD_ORDER_SUMMARY_COLUMNS = [
        "商品id",
        "订单笔数",
        "订单金额",
        "仅退款笔数",
        "仅退款金额",
        "实际发出笔数",
        "实际发出金额",
        "退货退款笔数",
        "退货退款金额",
        "实际成交笔数",
    ]
    TMALL_ORDER_ID_COLUMN_CANDIDATES = ("主订单编号", "订单编号", "订单号")
    TMALL_PRODUCT_ID_COLUMN_CANDIDATES = ("商品ID", "商品id", "商品Id", "宝贝ID", "宝贝id")
    TMALL_QUANTITY_COLUMN_CANDIDATES = ("购买数量", "数量", "商品数量", "购买件数")
    TMALL_AMOUNT_COLUMN_CANDIDATES = ("买家实付金额", "实付金额", "订单金额", "买家实付")
    TMALL_ORDER_STATUS_COLUMN_CANDIDATES = ("订单状态",)
    TMALL_LOGISTICS_COLUMN_CANDIDATES = ("物流单号", "运单号", "快递单号")
    TMALL_FAILED_STATUS_VALUES = ("交易失败", "交易关闭")
    TMALL_MONEY_SUMMARY_COLUMNS = (
        "订单金额",
        "仅退款金额",
        "实际发出金额",
        "退货退款金额",
    )
    TMALL_SUMMARY_BLANK_COLUMNS_AFTER_PRODUCT_ID = 2
    DOUYIN_ORDER_AMOUNT_COLUMN_CANDIDATES = (
        "订单应付金额",
        "买家实付金额",
        "实付金额",
        "订单金额",
        "买家实付",
    )
    DOUYIN_ORDER_STATUS_COLUMN_CANDIDATES = ("订单状态",)
    DOUYIN_SHIPMENT_TIME_COLUMN_CANDIDATES = ("发货时间", "物流单号", "运单号", "快递单号")
    DOUYIN_FAILED_STATUS_VALUES = ("已关闭", "交易关闭", "交易失败")

    def __init__(self) -> None:
        self.raw_df: Optional[pd.DataFrame] = None
        self.processed_df: Optional[pd.DataFrame] = None
        self.latest_file_path: Optional[Path] = None
        self.date_column: Optional[str] = None
        self.amount_column: Optional[str] = None
        self.refund_type_column: Optional[str] = None
        self.goods_status_column: Optional[str] = None

    def _build_empty_summary(self) -> dict[str, Any]:
        """
        构建空统计结果。
        """
        categories: dict[str, Any] = {}
        for category, sub_categories in self.CATEGORY_STRUCTURE.items():
            categories[category] = {
                "count": 0,
                "amount": 0.0,
                "sub_categories": {
                    sub_name: {"count": 0, "amount": 0.0} for sub_name in sub_categories
                },
            }

        return {
            "total_count": 0,
            "total_amount": 0.0,
            "categories": categories,
        }

    def _build_zero_douyin_refund_summary(self, report_date: str = "") -> dict[str, Any]:
        """
        构建抖音退款分析无数据时的 0 汇总。
        """
        summary = self._build_empty_summary()
        summary["douyin_refund_metrics"] = dict(self.DOUYIN_ZERO_REFUND_METRICS)
        if report_date:
            summary["report_date"] = report_date
        return summary

    @staticmethod
    def _extract_report_date_from_title(title: str) -> str:
        """
        从抖音退款明细标题/文件名中提取日期，支持 2026_05_13 / 2026-05-13 / 2026年05月13日。
        """
        text = str(title or "")
        patterns = (
            r"(20\d{2})[_\-.年/](\d{1,2})[_\-.月/](\d{1,2})",
            r"(20\d{2})(\d{2})(\d{2})",
        )
        for pattern in patterns:
            for match in re.finditer(pattern, text):
                try:
                    parsed = date(
                        int(match.group(1)),
                        int(match.group(2)),
                        int(match.group(3)),
                    )
                except ValueError:
                    continue
                return parsed.strftime("%Y-%m-%d")
        return ""

    def _find_column(self, df: pd.DataFrame, candidates: tuple[str, ...]) -> Optional[str]:
        """
        在候选列名中查找实际存在的列。
        """
        for name in candidates:
            if name in df.columns:
                return name
        return None

    def _resolve_excel_file(self, input_path: Path) -> Path:
        """
        解析输入路径，并自动定位最新 Excel 文件。
        """
        if input_path.is_file():
            if input_path.suffix.lower() not in self.SUPPORTED_EXCEL_SUFFIXES:
                raise ValueError(f"不支持的文件格式：{input_path.suffix}")
            return input_path
        return get_latest_file(
            directory=input_path,
            suffixes=self.SUPPORTED_EXCEL_SUFFIXES,
            ignore_prefixes=("~$",),
            recursive=False,
        )

    def load_data(self, file_path: Path) -> pd.DataFrame:
        """
        读取原始数据。
        支持传入目录（自动读取最新 Excel）或单个 Excel 文件。
        """
        target_file = self._resolve_excel_file(Path(file_path))
        self.latest_file_path = target_file
        df = pd.read_excel(target_file)
        self.raw_df = df
        return df

    def validate_columns(self, df: pd.DataFrame) -> None:
        """
        校验必要字段。
        """
        self.date_column = self._find_column(df, self.DATE_COLUMN_CANDIDATES)
        self.amount_column = self._find_column(df, self.AMOUNT_COLUMN_CANDIDATES)
        self.refund_type_column = self._find_column(df, self.REFUND_TYPE_COLUMN_CANDIDATES)
        self.goods_status_column = self._find_column(df, self.GOODS_STATUS_COLUMN_CANDIDATES)

        if self.date_column is None:
            candidates_text = "、".join(self.DATE_COLUMN_CANDIDATES)
            raise ValueError(f"未找到日期列，候选列名：{candidates_text}")

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        清洗原始数据。
        """
        cleaned = df.copy()
        cleaned.columns = [str(column).strip() for column in cleaned.columns]

        object_columns = cleaned.select_dtypes(include=["object"]).columns
        for column in object_columns:
            cleaned[column] = cleaned[column].where(
                cleaned[column].isna(),
                cleaned[column].astype(str).str.strip(),
            )

        return cleaned

    def _parse_amount(self, series: pd.Series) -> pd.Series:
        """
        将金额列标准化为浮点数。
        """
        cleaned = (
            series.fillna("")
            .astype(str)
            .str.replace(r"[^\d\.\-]", "", regex=True)
            .replace("", pd.NA)
        )
        return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)

    def _classify_row(self, refund_type: str, goods_status: str) -> tuple[Optional[str], Optional[str]]:
        """
        按分类规则识别一级/二级分类。
        """
        refund_text = str(refund_type or "").strip()
        status_text = str(goods_status or "").strip()
        combined_text = f"{refund_text}|{status_text}"

        # 按用户业务口径：优先由“货物状态”直接归类
        if "未发货" in status_text:
            return "未发货仅退款", "未发货"
        if "已寄回" in status_text:
            return "退货退款", "已寄回"
        if "未收到货" in status_text:
            return "已发货仅退款", "未收到货"
        if "已收到货" in status_text:
            return "已发货仅退款", "已收到货"

        if "未发货仅退款" in combined_text or "未发货" == status_text:
            return "未发货仅退款", "未发货"

        if "已发货仅退款" in combined_text:
            if "已收到货" in combined_text:
                return "已发货仅退款", "已收到货"
            if "未收到货" in combined_text:
                return "已发货仅退款", "未收到货"

        if "退货退款" in combined_text and "已寄回" in combined_text:
            return "退货退款", "已寄回"

        if "仅退款" in refund_text:
            if "未发货" in status_text:
                return "未发货仅退款", "未发货"
            if "已收到货" in status_text:
                return "已发货仅退款", "已收到货"
            if "未收到货" in status_text:
                return "已发货仅退款", "未收到货"

        if "已寄回" in status_text:
            return "退货退款", "已寄回"

        return None, None

    def transform_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        转换数据结构并筛选前一天数据。
        """
        return self.transform_data_for_date(df=df, report_date=None)

    def transform_data_for_date(
        self,
        df: pd.DataFrame,
        report_date: date | datetime | str | None = None,
    ) -> pd.DataFrame:
        """
        转换数据结构并筛选指定报表日期；未指定时默认前一天。
        """
        transformed = df.copy()

        transformed["_parsed_date"] = pd.to_datetime(
            transformed[self.date_column],
            errors="coerce",
        )
        target_date = self._normalize_report_date(report_date)
        transformed = transformed.loc[transformed["_parsed_date"].dt.date == target_date].copy()

        if transformed.empty:
            transformed["_amount"] = pd.Series(dtype="float64")
            transformed["_category"] = pd.Series(dtype="object")
            transformed["_sub_category"] = pd.Series(dtype="object")
            return transformed

        if self.amount_column and self.amount_column in transformed.columns:
            transformed["_amount"] = self._parse_amount(transformed[self.amount_column])
        else:
            transformed["_amount"] = 0.0

        if self.refund_type_column and self.refund_type_column in transformed.columns:
            refund_series = transformed[self.refund_type_column].fillna("").astype(str)
        else:
            refund_series = pd.Series("", index=transformed.index)

        if self.goods_status_column and self.goods_status_column in transformed.columns:
            status_series = transformed[self.goods_status_column].fillna("").astype(str)
        else:
            status_series = pd.Series("", index=transformed.index)

        category_pairs = [
            self._classify_row(refund_type=refund, goods_status=status)
            for refund, status in zip(refund_series, status_series)
        ]
        transformed["_category"] = [pair[0] for pair in category_pairs]
        transformed["_sub_category"] = [pair[1] for pair in category_pairs]
        return transformed

    @staticmethod
    def _normalize_report_date(report_date: date | datetime | str | None = None) -> date:
        """
        将外部传入的报表日期统一为 date；未传时沿用默认“昨天”。
        """
        if report_date is None:
            return get_previous_date()
        if isinstance(report_date, datetime):
            return report_date.date()
        if isinstance(report_date, date):
            return report_date
        parsed = datetime.strptime(str(report_date).strip(), "%Y-%m-%d")
        return parsed.date()

    def summarize_data(self, df: pd.DataFrame) -> dict[str, Any]:
        """
        汇总统计数据，返回字典结果。
        """
        summary = self._build_empty_summary()
        if df.empty:
            return summary

        summary["total_count"] = int(len(df))
        summary["total_amount"] = round(float(df["_amount"].sum()), 2)

        for category_name, sub_categories in self.CATEGORY_STRUCTURE.items():
            category_df = df.loc[df["_category"] == category_name]
            summary["categories"][category_name]["count"] = int(len(category_df))
            summary["categories"][category_name]["amount"] = round(
                float(category_df["_amount"].sum()),
                2,
            )

            for sub_name in sub_categories:
                sub_df = category_df.loc[category_df["_sub_category"] == sub_name]
                summary["categories"][category_name]["sub_categories"][sub_name]["count"] = int(
                    len(sub_df)
                )
                summary["categories"][category_name]["sub_categories"][sub_name]["amount"] = round(
                    float(sub_df["_amount"].sum()),
                    2,
                )

        return summary

    @staticmethod
    def _normalize_column_name(name: Any) -> str:
        """
        标准化列名，去掉多余空白。
        """
        return re.sub(r"\s+", "", str(name or "")).strip()

    def _find_normalized_column(self, df: pd.DataFrame, target_name: str) -> Optional[str]:
        """
        按标准化列名查找真实列名。
        """
        normalized_target = self._normalize_column_name(target_name)
        for column in df.columns:
            if self._normalize_column_name(column) == normalized_target:
                return str(column)
        return None

    def _find_first_normalized_column(
        self,
        df: pd.DataFrame,
        candidates: tuple[str, ...],
    ) -> Optional[str]:
        """
        在多个候选列名中按标准化列名查找真实列名。
        """
        for candidate in candidates:
            column = self._find_normalized_column(df, candidate)
            if column is not None:
                return column
        return None

    def _require_first_normalized_column(
        self,
        df: pd.DataFrame,
        candidates: tuple[str, ...],
        usage: str,
    ) -> str:
        """
        查找必要列，缺失时给出业务可读错误。
        """
        column = self._find_first_normalized_column(df, candidates)
        if column is None:
            raise ValueError(f"天猫订单表缺少{usage}列，候选列名：{'、'.join(candidates)}")
        return column

    @staticmethod
    def _normalize_identifier(value: Any) -> str:
        """
        标准化订单号、商品 ID、物流单号等标识，避免 Excel 数值读入后带 .0。
        """
        if pd.isna(value):
            return ""
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        if re.fullmatch(r"\d+\.0", text):
            return text[:-2]
        return text

    @classmethod
    def _normalize_optional_identifier(cls, value: Any) -> str:
        """
        标准化可为空字段，把抖音 CSV 中的占位符视为空。
        """
        text = cls._normalize_identifier(value)
        if text.strip().lower() in {"-", "无", "nan", "none", "null"}:
            return ""
        return text

    @classmethod
    def _normalize_identifier_list(
        cls,
        values: str | tuple[str, ...] | list[str] | None,
    ) -> tuple[str, ...]:
        """
        标准化商品 ID 列表，支持逗号、中文逗号、分号、空白和换行。
        """
        if values is None:
            return ()
        if isinstance(values, (tuple, list)):
            raw_items = [str(item or "") for item in values]
        else:
            raw_items = re.split(r"[,，;；\s]+", str(values or ""))

        normalized: list[str] = []
        for item in raw_items:
            token = cls._normalize_identifier(item)
            if token and token not in normalized:
                normalized.append(token)
        return tuple(normalized)

    @classmethod
    def _unique_identifier_count(cls, series: pd.Series) -> int:
        values = [cls._normalize_identifier(value) for value in series]
        return len({value for value in values if value})

    def _sum_numeric_column(self, df: pd.DataFrame, column_name: str) -> float:
        """
        汇总指定列的数值，兼容金额符号与千分位。
        """
        if column_name not in df.columns:
            return 0.0
        values = self._parse_amount(df[column_name])
        return round(float(values.sum()), 2)

    def _sum_count_column(self, df: pd.DataFrame, column_name: str) -> int:
        """
        汇总指定列的订单数。
        """
        return int(round(self._sum_numeric_column(df=df, column_name=column_name)))

    def _read_table_file(self, input_path: Path) -> pd.DataFrame:
        """
        读取 Excel/CSV 表格，CSV 优先按 UTF-8 BOM 解析。
        """
        target_file = Path(input_path)
        suffix = target_file.suffix.lower()
        if suffix == ".csv":
            for encoding in ("utf-8-sig", "utf-8", "gb18030"):
                try:
                    return pd.read_csv(target_file, encoding=encoding, dtype=str)
                except UnicodeDecodeError:
                    continue
            return pd.read_csv(target_file, dtype=str)
        return pd.read_excel(target_file)

    def summarize_douyin_refund_analysis(self, input_path: Path) -> dict[str, Any]:
        """
        汇总抖音电商罗盘“退款分析 -> 下载明细”文件中的本店数据。
        """
        target_file = self._resolve_excel_file(Path(input_path))
        excel_file = pd.ExcelFile(target_file)
        if self.DOUYIN_REFUND_DETAIL_SHEET_NAME not in excel_file.sheet_names:
            sheets = "、".join(excel_file.sheet_names)
            raise ValueError(
                f"未找到抖音退款分析工作表【{self.DOUYIN_REFUND_DETAIL_SHEET_NAME}】，当前工作表：{sheets}"
            )

        df = pd.read_excel(target_file, sheet_name=self.DOUYIN_REFUND_DETAIL_SHEET_NAME)
        df.columns = [str(column).strip() for column in df.columns]
        report_date = self._extract_report_date_from_title(target_file.stem)
        date_column = self._find_normalized_column(df, "日期")
        if not report_date and date_column is not None:
            parsed_dates = pd.to_datetime(df[date_column], errors="coerce")
            valid_dates = parsed_dates.dropna()
            if not valid_dates.empty:
                report_date = valid_dates.dt.date.max().strftime("%Y-%m-%d")

        resolved_columns: dict[str, str] = {}
        missing_columns: list[str] = []
        for metric_key, expected_column in self.DOUYIN_REFUND_METRIC_COLUMNS.items():
            actual_column = self._find_normalized_column(df, expected_column)
            if actual_column is None:
                missing_columns.append(expected_column)
            else:
                resolved_columns[metric_key] = actual_column

        if missing_columns:
            all_metric_columns_missing = len(missing_columns) == len(self.DOUYIN_REFUND_METRIC_COLUMNS)
            if all_metric_columns_missing or df.empty:
                return self._build_zero_douyin_refund_summary(report_date=report_date)
            raise ValueError(f"抖音退款分析明细缺少列：{'、'.join(missing_columns)}")

        metrics = {
            "refund_total_order_count": self._sum_count_column(
                df, resolved_columns["refund_total_order_count"]
            ),
            "refund_total_amount": self._sum_numeric_column(
                df, resolved_columns["refund_total_amount"]
            ),
            "pre_shipment_refund_order_count": self._sum_count_column(
                df, resolved_columns["pre_shipment_refund_order_count"]
            ),
            "pre_shipment_refund_amount": self._sum_numeric_column(
                df, resolved_columns["pre_shipment_refund_amount"]
            ),
            "unreceived_refund_order_count": self._sum_count_column(
                df, resolved_columns["unreceived_refund_order_count"]
            ),
            "unreceived_refund_amount": self._sum_numeric_column(
                df, resolved_columns["unreceived_refund_amount"]
            ),
            "received_refund_order_count": self._sum_count_column(
                df, resolved_columns["received_refund_order_count"]
            ),
            "received_refund_amount": self._sum_numeric_column(
                df, resolved_columns["received_refund_amount"]
            ),
            "return_refund_order_count": self._sum_count_column(
                df, resolved_columns["return_refund_order_count"]
            ),
            "return_refund_amount": self._sum_numeric_column(
                df, resolved_columns["return_refund_amount"]
            ),
        }

        summary = self._build_empty_summary()
        summary["total_count"] = metrics["refund_total_order_count"]
        summary["total_amount"] = metrics["refund_total_amount"]
        summary["douyin_refund_metrics"] = metrics
        if report_date:
            summary["report_date"] = report_date

        summary["categories"]["未发货仅退款"]["count"] = metrics[
            "pre_shipment_refund_order_count"
        ]
        summary["categories"]["未发货仅退款"]["amount"] = metrics["pre_shipment_refund_amount"]
        summary["categories"]["未发货仅退款"]["sub_categories"]["未发货"]["count"] = metrics[
            "pre_shipment_refund_order_count"
        ]
        summary["categories"]["未发货仅退款"]["sub_categories"]["未发货"]["amount"] = metrics[
            "pre_shipment_refund_amount"
        ]

        shipped_count = (
            metrics["unreceived_refund_order_count"] + metrics["received_refund_order_count"]
        )
        shipped_amount = round(
            metrics["unreceived_refund_amount"] + metrics["received_refund_amount"],
            2,
        )
        summary["categories"]["已发货仅退款"]["count"] = shipped_count
        summary["categories"]["已发货仅退款"]["amount"] = shipped_amount
        summary["categories"]["已发货仅退款"]["sub_categories"]["未收到货"]["count"] = metrics[
            "unreceived_refund_order_count"
        ]
        summary["categories"]["已发货仅退款"]["sub_categories"]["未收到货"]["amount"] = metrics[
            "unreceived_refund_amount"
        ]
        summary["categories"]["已发货仅退款"]["sub_categories"]["已收到货"]["count"] = metrics[
            "received_refund_order_count"
        ]
        summary["categories"]["已发货仅退款"]["sub_categories"]["已收到货"]["amount"] = metrics[
            "received_refund_amount"
        ]

        summary["categories"]["退货退款"]["count"] = metrics["return_refund_order_count"]
        summary["categories"]["退货退款"]["amount"] = metrics["return_refund_amount"]
        summary["categories"]["退货退款"]["sub_categories"]["已寄回"]["count"] = metrics[
            "return_refund_order_count"
        ]
        summary["categories"]["退货退款"]["sub_categories"]["已寄回"]["amount"] = metrics[
            "return_refund_amount"
        ]
        return summary

    def summarize_douyin_after_sale_orders(self, input_path: Path) -> dict[str, Any]:
        """
        汇总抖店“售后工作台 -> 导出”的售后单。
        """
        target_file = Path(input_path)
        if target_file.suffix.lower() not in self.SUPPORTED_TABLE_SUFFIXES:
            target_file = self._resolve_excel_file(target_file)
        df = self._read_table_file(target_file)
        df.columns = [str(column).strip() for column in df.columns]

        type_column = self._find_normalized_column(df, self.DOUYIN_AFTER_SALE_TYPE_COLUMN)
        amount_column = self._find_normalized_column(df, self.DOUYIN_AFTER_SALE_AMOUNT_COLUMN)
        if type_column is None or amount_column is None:
            missing_columns = []
            if type_column is None:
                missing_columns.append(self.DOUYIN_AFTER_SALE_TYPE_COLUMN)
            if amount_column is None:
                missing_columns.append(self.DOUYIN_AFTER_SALE_AMOUNT_COLUMN)
            raise ValueError(f"抖音售后单缺少列：{'、'.join(missing_columns)}")

        finished_time_column = self._find_normalized_column(
            df, self.DOUYIN_AFTER_SALE_FINISHED_TIME_COLUMN
        )
        report_date = get_previous_date().strftime("%Y-%m-%d")
        if finished_time_column is not None:
            parsed_dates = pd.to_datetime(df[finished_time_column], errors="coerce")
            valid_dates = parsed_dates.dropna()
            if not valid_dates.empty:
                report_date = valid_dates.dt.date.max().strftime("%Y-%m-%d")

        type_series = df[type_column].fillna("").astype(str).str.strip()
        amount_series = self._parse_amount(df[amount_column])

        def count_amount(type_name: str) -> tuple[int, float]:
            mask = type_series == type_name
            return int(mask.sum()), round(float(amount_series.loc[mask].sum()), 2)

        return_count, return_amount = count_amount("退货退款")
        shipped_count, shipped_amount = count_amount("已发货退款")
        unshipped_count, unshipped_amount = count_amount("未发货退款")
        total_count = int(return_count + shipped_count + unshipped_count)
        total_amount = round(float(return_amount + shipped_amount + unshipped_amount), 2)

        metrics = {
            "refund_total_order_count": total_count,
            "refund_total_amount": total_amount,
            "pre_shipment_refund_order_count": unshipped_count,
            "pre_shipment_refund_amount": unshipped_amount,
            "unreceived_refund_order_count": 0,
            "unreceived_refund_amount": 0.0,
            "received_refund_order_count": shipped_count,
            "received_refund_amount": shipped_amount,
            "return_refund_order_count": return_count,
            "return_refund_amount": return_amount,
        }

        summary = self._build_empty_summary()
        summary["total_count"] = total_count
        summary["total_amount"] = total_amount
        summary["douyin_refund_metrics"] = metrics
        if report_date:
            summary["report_date"] = report_date

        summary["categories"]["未发货仅退款"]["count"] = unshipped_count
        summary["categories"]["未发货仅退款"]["amount"] = unshipped_amount
        summary["categories"]["未发货仅退款"]["sub_categories"]["未发货"]["count"] = unshipped_count
        summary["categories"]["未发货仅退款"]["sub_categories"]["未发货"]["amount"] = unshipped_amount

        summary["categories"]["已发货仅退款"]["count"] = shipped_count
        summary["categories"]["已发货仅退款"]["amount"] = shipped_amount
        summary["categories"]["已发货仅退款"]["sub_categories"]["已收到货"]["count"] = shipped_count
        summary["categories"]["已发货仅退款"]["sub_categories"]["已收到货"]["amount"] = shipped_amount

        summary["categories"]["退货退款"]["count"] = return_count
        summary["categories"]["退货退款"]["amount"] = return_amount
        summary["categories"]["退货退款"]["sub_categories"]["已寄回"]["count"] = return_count
        summary["categories"]["退货退款"]["sub_categories"]["已寄回"]["amount"] = return_amount
        return summary

    def summarize_tmall_sold_orders(
        self,
        input_path: Path,
        product_ids: str | tuple[str, ...] | list[str] | None = None,
    ) -> pd.DataFrame:
        """
        汇总天猫/淘宝【已卖出宝贝】宝贝销售明细报表。
        """
        target_file = Path(input_path)
        if target_file.suffix.lower() not in self.SUPPORTED_TABLE_SUFFIXES:
            target_file = self._resolve_excel_file(target_file)
        df = self._read_table_file(target_file)
        df = self.clean_data(df)

        if df.empty:
            return pd.DataFrame(columns=self.TMALL_SOLD_ORDER_SUMMARY_COLUMNS)

        order_id_column = self._require_first_normalized_column(
            df,
            self.TMALL_ORDER_ID_COLUMN_CANDIDATES,
            "订单号",
        )
        product_id_column = self._require_first_normalized_column(
            df,
            self.TMALL_PRODUCT_ID_COLUMN_CANDIDATES,
            "商品 ID",
        )
        quantity_column = self._require_first_normalized_column(
            df,
            self.TMALL_QUANTITY_COLUMN_CANDIDATES,
            "购买数量",
        )
        amount_column = self._require_first_normalized_column(
            df,
            self.TMALL_AMOUNT_COLUMN_CANDIDATES,
            "实付金额",
        )
        status_column = self._require_first_normalized_column(
            df,
            self.TMALL_ORDER_STATUS_COLUMN_CANDIDATES,
            "订单状态",
        )
        logistics_column = self._require_first_normalized_column(
            df,
            self.TMALL_LOGISTICS_COLUMN_CANDIDATES,
            "物流单号",
        )

        working = pd.DataFrame(
            {
                "_product_id": df[product_id_column].map(self._normalize_identifier),
                "_order_id": df[order_id_column].map(self._normalize_identifier),
                "_quantity": self._parse_amount(df[quantity_column]),
                "_amount": self._parse_amount(df[amount_column]),
                "_status": df[status_column].fillna("").astype(str).str.strip(),
                "_logistics": df[logistics_column].map(self._normalize_identifier),
            }
        )
        working = working.loc[working["_product_id"] != ""].copy()
        requested_product_ids = self._normalize_identifier_list(product_ids)
        if requested_product_ids:
            requested_set = set(requested_product_ids)
            working = working.loc[working["_product_id"].isin(requested_set)].copy()
        if working.empty:
            return pd.DataFrame(columns=self.TMALL_SOLD_ORDER_SUMMARY_COLUMNS)

        rows: list[dict[str, Any]] = []
        grouped_products = {
            product_id: product_df
            for product_id, product_df in working.groupby("_product_id", sort=False)
        }
        product_order = requested_product_ids if requested_product_ids else tuple(grouped_products.keys())
        for product_id in product_order:
            product_df = grouped_products.get(product_id)
            if product_df is None:
                continue
            failed_mask = product_df["_status"].map(
                lambda value: any(status in str(value) for status in self.TMALL_FAILED_STATUS_VALUES)
            )
            failed_df = product_df.loc[failed_mask]
            only_refund_df = failed_df.loc[failed_df["_logistics"] == ""]
            return_refund_df = failed_df.loc[failed_df["_logistics"] != ""]

            order_count = int(len(product_df))
            order_amount = round(float(product_df["_amount"].sum()), 2)
            only_refund_count = int(len(only_refund_df))
            only_refund_amount = round(float(only_refund_df["_amount"].sum()), 2)
            return_refund_count = int(len(return_refund_df))
            return_refund_amount = round(float(return_refund_df["_amount"].sum()), 2)

            actual_sent_count = max(order_count - only_refund_count, 0)
            actual_sent_amount = round(order_amount - only_refund_amount, 2)
            actual_deal_count = max(order_count - only_refund_count - return_refund_count, 0)

            rows.append(
                {
                    "商品id": product_id,
                    "订单笔数": order_count,
                    "订单金额": order_amount,
                    "仅退款笔数": only_refund_count,
                    "仅退款金额": only_refund_amount,
                    "实际发出笔数": actual_sent_count,
                    "实际发出金额": actual_sent_amount,
                    "退货退款笔数": return_refund_count,
                    "退货退款金额": return_refund_amount,
                    "实际成交笔数": actual_deal_count,
                }
            )

        return pd.DataFrame(rows, columns=self.TMALL_SOLD_ORDER_SUMMARY_COLUMNS)

    def summarize_douyin_order_details(
        self,
        input_path: Path,
        product_ids: str | tuple[str, ...] | list[str] | None = None,
    ) -> pd.DataFrame:
        """
        汇总抖音订单明细离线导出表。
        """
        target_file = Path(input_path)
        if target_file.suffix.lower() not in self.SUPPORTED_TABLE_SUFFIXES:
            target_file = self._resolve_excel_file(target_file)
        df = self._read_table_file(target_file)
        df = self.clean_data(df)

        if df.empty:
            return pd.DataFrame(columns=self.TMALL_SOLD_ORDER_SUMMARY_COLUMNS)

        product_id_column = self._require_first_normalized_column(
            df,
            self.TMALL_PRODUCT_ID_COLUMN_CANDIDATES,
            "商品 ID",
        )
        quantity_column = self._require_first_normalized_column(
            df,
            self.TMALL_QUANTITY_COLUMN_CANDIDATES,
            "购买数量",
        )
        amount_column = self._require_first_normalized_column(
            df,
            self.DOUYIN_ORDER_AMOUNT_COLUMN_CANDIDATES,
            "实付金额",
        )
        status_column = self._require_first_normalized_column(
            df,
            self.DOUYIN_ORDER_STATUS_COLUMN_CANDIDATES,
            "订单状态",
        )
        shipment_time_column = self._require_first_normalized_column(
            df,
            self.DOUYIN_SHIPMENT_TIME_COLUMN_CANDIDATES,
            "发货时间",
        )

        working = pd.DataFrame(
            {
                "_product_id": df[product_id_column].map(self._normalize_identifier),
                "_quantity": self._parse_amount(df[quantity_column]),
                "_amount": self._parse_amount(df[amount_column]),
                "_status": df[status_column].fillna("").astype(str).str.strip(),
                "_shipment_time": df[shipment_time_column].map(self._normalize_optional_identifier),
            }
        )
        working = working.loc[working["_product_id"] != ""].copy()
        requested_product_ids = self._normalize_identifier_list(product_ids)
        if requested_product_ids:
            requested_set = set(requested_product_ids)
            working = working.loc[working["_product_id"].isin(requested_set)].copy()
        if working.empty:
            return pd.DataFrame(columns=self.TMALL_SOLD_ORDER_SUMMARY_COLUMNS)

        rows: list[dict[str, Any]] = []
        grouped_products = {
            product_id: product_df
            for product_id, product_df in working.groupby("_product_id", sort=False)
        }
        if requested_product_ids:
            product_order = requested_product_ids
        else:
            product_order = tuple(
                product_id
                for product_id, _product_df in sorted(
                    grouped_products.items(),
                    key=lambda item: float(item[1]["_quantity"].sum()),
                    reverse=True,
                )
            )
        for product_id in product_order:
            product_df = grouped_products.get(product_id)
            if product_df is None:
                continue
            refunded_mask = product_df["_status"].map(
                lambda value: any(status in str(value) for status in self.DOUYIN_FAILED_STATUS_VALUES)
            )
            refunded_df = product_df.loc[refunded_mask]
            only_refund_df = refunded_df.loc[refunded_df["_shipment_time"] == ""]
            return_refund_df = refunded_df.loc[refunded_df["_shipment_time"] != ""]

            order_count = int(round(float(product_df["_quantity"].sum())))
            order_amount = round(float(product_df["_amount"].sum()), 2)
            only_refund_count = int(round(float(only_refund_df["_quantity"].sum())))
            only_refund_amount = round(float(only_refund_df["_amount"].sum()), 2)
            return_refund_count = int(round(float(return_refund_df["_quantity"].sum())))
            return_refund_amount = round(float(return_refund_df["_amount"].sum()), 2)

            actual_sent_count = max(order_count - only_refund_count, 0)
            actual_sent_amount = round(order_amount - only_refund_amount, 2)
            actual_deal_count = max(order_count - only_refund_count - return_refund_count, 0)

            rows.append(
                {
                    "商品id": product_id,
                    "订单笔数": order_count,
                    "订单金额": order_amount,
                    "仅退款笔数": only_refund_count,
                    "仅退款金额": only_refund_amount,
                    "实际发出笔数": actual_sent_count,
                    "实际发出金额": actual_sent_amount,
                    "退货退款笔数": return_refund_count,
                    "退货退款金额": return_refund_amount,
                    "实际成交笔数": actual_deal_count,
                }
            )

        return pd.DataFrame(rows, columns=self.TMALL_SOLD_ORDER_SUMMARY_COLUMNS)

    def save_tmall_sold_order_summary(
        self,
        input_path: Path,
        output_path: Path,
        product_ids: str | tuple[str, ...] | list[str] | None = None,
    ) -> Path:
        """
        生成天猫/淘宝【已卖出宝贝】订单汇总表。
        """
        summary_df = self.summarize_tmall_sold_orders(input_path, product_ids=product_ids)
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        summary_df.to_excel(output_file, index=False)
        self._format_tmall_sold_order_summary(output_file)
        return output_file

    def save_douyin_order_detail_summary(
        self,
        input_path: Path,
        output_path: Path,
        product_ids: str | tuple[str, ...] | list[str] | None = None,
    ) -> Path:
        """
        生成抖音订单明细离线汇总表。
        """
        summary_df = self.summarize_douyin_order_details(input_path, product_ids=product_ids)
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        summary_df.to_excel(output_file, index=False)
        self._format_tmall_sold_order_summary(output_file)
        return output_file

    def _format_tmall_sold_order_summary(self, output_file: Path) -> None:
        """
        设置汇总表金额列格式，确保 Excel 中保留两位小数显示。
        """
        from openpyxl import load_workbook

        workbook = load_workbook(output_file)
        worksheet = workbook.active
        if worksheet.max_column >= 2:
            worksheet.insert_cols(2, amount=self.TMALL_SUMMARY_BLANK_COLUMNS_AFTER_PRODUCT_ID)
        header_by_column = {
            cell.column: str(cell.value or "").strip()
            for cell in worksheet[1]
        }
        for column_index in range(1, worksheet.max_column + 1):
            header = header_by_column.get(column_index, "")
            column_letter = worksheet.cell(row=1, column=column_index).column_letter
            values = [
                worksheet.cell(row=row_index, column=column_index).value
                for row_index in range(1, worksheet.max_row + 1)
            ]
            max_text_length = max((len(str(value)) for value in values if value is not None), default=0)
            if header == "商品id":
                width = max(max_text_length + 2, 22)
            elif header in self.TMALL_MONEY_SUMMARY_COLUMNS:
                width = max(max_text_length + 4, 14)
            elif header:
                width = max(max_text_length + 2, 12)
            else:
                width = 10
            worksheet.column_dimensions[column_letter].width = min(width, 28)

        for column_index, header in header_by_column.items():
            if header not in self.TMALL_MONEY_SUMMARY_COLUMNS:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = "0.00"
        workbook.save(output_file)

    def save_processed_data(self, df: pd.DataFrame, output_path: Path) -> None:
        """
        保存处理后的中间结果。
        """
        export_df = df.copy()
        if "_category" in export_df.columns:
            export_df["一级分类"] = export_df["_category"]
        if "_sub_category" in export_df.columns:
            export_df["二级分类"] = export_df["_sub_category"]
        if "_amount" in export_df.columns:
            export_df["标准化金额"] = export_df["_amount"]

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        export_df.to_excel(output_path, index=False)

    def process(
        self,
        input_path: Path,
        output_path: Optional[Path] = None,
        report_date: date | datetime | str | None = None,
    ) -> dict[str, Any]:
        """
        执行完整数据处理流程。
        """
        loaded_df = self.load_data(Path(input_path))
        cleaned_df = self.clean_data(loaded_df)
        if cleaned_df.empty:
            self.processed_df = cleaned_df
            if output_path is not None:
                self.save_processed_data(cleaned_df, Path(output_path))
            return self._build_empty_summary()

        self.validate_columns(cleaned_df)
        transformed_df = self.transform_data_for_date(cleaned_df, report_date=report_date)
        self.processed_df = transformed_df

        if output_path is not None:
            self.save_processed_data(transformed_df, Path(output_path))

        return self.summarize_data(transformed_df)
