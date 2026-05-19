"""
Excel 输出模块。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from qianiu_auto_report.config import (
    DateConfig,
    EXCEL_OUTPUT_DIR,
    REPORT_TEMPLATE_PATH,
    TEMPLATE_DIR,
)


class ExcelWriter:
    """
    Excel 报表写入器。
    """

    PLATFORM_FILENAME_PREFIX = {
        "taobao": "淘宝",
        "douyin": "抖音",
    }

    SUMMARY_CELL_MAPPING = {
        "total_count": "B2",
        "total_amount": "C2",
        "已发货仅退款.count": "B4",
        "已发货仅退款.amount": "C4",
        "已发货仅退款.未收到货.count": "B5",
        "已发货仅退款.未收到货.amount": "C5",
        "已发货仅退款.已收到货.count": "B6",
        "已发货仅退款.已收到货.amount": "C6",
        "未发货仅退款.count": "B8",
        "未发货仅退款.amount": "C8",
        "未发货仅退款.未发货.count": "B9",
        "未发货仅退款.未发货.amount": "C9",
        "退货退款.count": "B11",
        "退货退款.amount": "C11",
        "退货退款.已寄回.count": "B12",
        "退货退款.已寄回.amount": "C12",
    }

    def __init__(self, template_path: Optional[Path] = None) -> None:
        self.template_path = template_path
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None

    def _resolve_template_path(self) -> Path:
        """
        解析模板路径，优先使用 `template.xlsx`。
        """
        if self.template_path is not None:
            resolved = Path(self.template_path)
            if not resolved.exists():
                raise FileNotFoundError(f"模板文件不存在：{resolved}")
            return resolved

        candidates = [
            TEMPLATE_DIR / "template.xlsx",
            REPORT_TEMPLATE_PATH,
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate

        searched = "、".join(str(path) for path in candidates)
        raise FileNotFoundError(f"未找到模板文件，请准备 template.xlsx。已检查路径：{searched}")

    def create_workbook(self) -> None:
        """
        创建工作簿（当模板不存在时可作为扩展入口）。
        """
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def load_template(self) -> None:
        """
        加载 Excel 模板。
        """
        template_file = self._resolve_template_path()
        self.workbook = load_workbook(template_file)
        self.worksheet = self.workbook.active

    def prepare_sheet(self, sheet_name: str = "") -> None:
        """
        准备工作表，默认使用激活页。
        """
        if self.workbook is None:
            raise RuntimeError("工作簿未加载，请先调用 load_template。")

        if sheet_name and sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[sheet_name]
        elif self.worksheet is None:
            self.worksheet = self.workbook.active

    def _extract_value(self, stats: dict[str, Any], key_path: str) -> Any:
        """
        从统计字典中按路径提取值，不存在时返回 0。
        """
        if key_path == "total_count":
            return stats.get("total_count", 0)
        if key_path == "total_amount":
            return stats.get("total_amount", 0.0)

        categories = stats.get("categories", {})
        parts = key_path.split(".")
        if len(parts) == 2:
            category_name, metric = parts
            category_data = categories.get(category_name, {})
            return category_data.get(metric, 0 if metric == "count" else 0.0)

        if len(parts) == 3:
            category_name, sub_name, metric = parts
            category_data = categories.get(category_name, {})
            sub_data = category_data.get("sub_categories", {}).get(sub_name, {})
            return sub_data.get(metric, 0 if metric == "count" else 0.0)

        return 0

    def write_summary(self, stats: dict[str, Any]) -> None:
        """
        将统计结果写入指定单元格。
        """
        if self.worksheet is None:
            raise RuntimeError("工作表未初始化，请先调用 prepare_sheet。")

        for key_path, cell in self.SUMMARY_CELL_MAPPING.items():
            value = self._extract_value(stats, key_path)
            self.worksheet[cell] = value

            if key_path.endswith(".amount") or key_path == "total_amount":
                self.worksheet[cell].number_format = "0.00"

    @staticmethod
    def _sanitize_filename_token(name: str, fallback: str = "报表") -> str:
        """
        清洗文件名片段，移除系统不允许字符。
        """
        token = re.sub(r"[\\/:*?\"<>|\r\n\t]+", "_", str(name or "")).strip(" ._")
        token = re.sub(r"\s+", "", token)
        return token or fallback

    def _build_output_file_path(
        self,
        output_path: Optional[Path] = None,
        report_date: Optional[str] = None,
        shop_name: str = "",
        default_prefix: str = "报表",
    ) -> Path:
        """
        生成输出文件路径，默认命名为 `店铺名_YYYY-MM-DD.xlsx`，无店铺名时使用默认前缀。
        """
        date_str = str(report_date or DateConfig.default_report_date_str()).strip()
        prefix = self._sanitize_filename_token(shop_name, fallback=default_prefix)
        file_name = f"{prefix}_{date_str}.xlsx"

        if output_path is None:
            target_dir = EXCEL_OUTPUT_DIR
        else:
            path_obj = Path(output_path)
            target_dir = path_obj if path_obj.suffix == "" else path_obj.parent

        target_dir.mkdir(parents=True, exist_ok=True)
        return target_dir / file_name

    def _build_platform_prefixed_name(
        self,
        shop_name: str,
        platform: Any,
        default_prefix: str,
    ) -> str:
        """
        生成带平台前缀的报表名片段。
        """
        platform_key = str(platform or "").strip().lower()
        platform_prefix = self.PLATFORM_FILENAME_PREFIX.get(platform_key, "")
        base_name = str(shop_name or "").strip() or default_prefix
        if not platform_prefix:
            return base_name
        if base_name.startswith(f"{platform_prefix}_") or base_name.startswith(platform_prefix):
            return base_name
        return f"{platform_prefix}_{base_name}"

    def save(self, output_path: Optional[Path] = None) -> Path:
        """
        保存 Excel 文件，不修改原模板。
        """
        if self.workbook is None:
            raise RuntimeError("工作簿未初始化，无法保存。")

        target_file = self._build_output_file_path(output_path)
        self.workbook.save(target_file)
        return target_file

    def export(
        self,
        df: dict[str, Any],
        output_path: Optional[Path] = None,
        sheet_name: str = "",
    ) -> Path:
        """
        执行完整 Excel 导出流程并返回新文件路径。
        """
        self.load_template()
        self.prepare_sheet(sheet_name=sheet_name)
        self.write_summary(stats=df)
        return self.save(output_path=output_path)

    def export_business_finance_metrics(
        self,
        metrics: dict[str, Any],
        output_path: Optional[Path] = None,
    ) -> Path:
        """
        输出业务/财务提取结果为独立 Excel 报表。
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "业务财务汇总"
        self._write_business_finance_metrics_sheet(worksheet=worksheet, metrics=metrics)

        if output_path is None:
            target_dir = Path.home() / "Desktop"
        else:
            path_obj = Path(output_path)
            target_dir = path_obj if path_obj.suffix == "" else path_obj.parent
        report_date = str(metrics.get("report_date", DateConfig.default_report_date_str()))
        shop_name = self._build_platform_prefixed_name(
            shop_name=str(metrics.get("shop_name", "")).strip(),
            platform=metrics.get("platform", ""),
            default_prefix="业务财务汇总",
        )
        target_file = self._build_output_file_path(
            output_path=target_dir,
            report_date=report_date,
            shop_name=shop_name,
            default_prefix="业务财务汇总",
        )
        workbook.save(target_file)
        return target_file

    def _build_business_finance_headers_values(
        self,
        metrics: dict[str, Any],
        refund_summary: Optional[dict[str, Any]] = None,
    ) -> tuple[list[str], list[Any]]:
        """
        构建业务财务汇总的表头与数据行。
        """
        report_date = str(metrics.get("report_date", DateConfig.default_report_date_str()))
        if refund_summary is None:
            embedded_refund_summary = metrics.get("refund_summary")
            if isinstance(embedded_refund_summary, dict):
                refund_summary = embedded_refund_summary

        refund_fields = self._extract_output_refund_summary_fields(refund_summary or {})
        headers = [
            "统计日期",
            "支付买家数",
            "支付子订单数",
            "支付金额",
            "未发货仅退款",
            "未发货仅退款金额",
            "已发货仅退款数量",
            "已发货仅退款金额",
            "退货退款数量",
            "退货退款金额",
            "交易赔付",
            "淘宝天猫跨境服务增值费",
            "推广费用",
        ]
        values = [
            report_date,
            metrics.get("payment_buyer_count", 0),
            metrics.get("payment_sub_order_count", 0),
            metrics.get("payment_amount", 0.0),
            refund_fields["unshipped_only_refund_count"],
            refund_fields["unshipped_only_refund_amount"],
            refund_fields["shipped_only_refund_count"],
            refund_fields["shipped_only_refund_amount"],
            refund_fields["return_refund_count"],
            refund_fields["return_refund_amount"],
            metrics.get("trade_compensation", 0.0),
            metrics.get("cross_border_value_added_fee", 0.0),
            metrics.get("promotion_fee", 0.0),
        ]
        return headers, values

    def _write_business_finance_metrics_sheet(
        self,
        worksheet: Worksheet,
        metrics: dict[str, Any],
        refund_summary: Optional[dict[str, Any]] = None,
    ) -> None:
        """
        将业务财务指标写入指定工作表。
        """
        headers, values = self._build_business_finance_headers_values(
            metrics=metrics,
            refund_summary=refund_summary,
        )

        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column_index, value=header)
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=2, column=column_index, value=value)
            if "金额" in headers[column_index - 1] or headers[column_index - 1] in {
                "交易赔付",
                "淘宝天猫跨境服务增值费",
                "推广费用",
            }:
                cell.number_format = "0.00"
            if (
                "数量" in headers[column_index - 1]
                or "订单数" in headers[column_index - 1]
                or headers[column_index - 1]
                in {
                "支付买家数",
                "支付子订单数",
                }
            ):
                cell.number_format = "0"

        for column_index, header in enumerate(headers, start=1):
            column_letter = worksheet.cell(row=1, column=column_index).column_letter
            fallback_width = max(len(str(header)) + 2, 12)
            worksheet.column_dimensions[column_letter].width = fallback_width

    def _extract_output_refund_summary_fields(self, refund_summary: dict[str, Any]) -> dict[str, Any]:
        """
        提取最终横向报表中的三类退款字段，淘宝/抖音共用同一输出口径。
        """
        if isinstance(refund_summary.get("douyin_refund_metrics"), dict):
            refund_fields = self._extract_douyin_refund_summary_fields(refund_summary)
            return {
                "unshipped_only_refund_count": refund_fields["pre_shipment_refund_order_count"],
                "unshipped_only_refund_amount": refund_fields["pre_shipment_refund_amount"],
                "shipped_only_refund_count": (
                    refund_fields["unreceived_refund_order_count"]
                    + refund_fields["received_refund_order_count"]
                ),
                "shipped_only_refund_amount": round(
                    refund_fields["unreceived_refund_amount"]
                    + refund_fields["received_refund_amount"],
                    2,
                ),
                "return_refund_count": refund_fields["return_refund_order_count"],
                "return_refund_amount": refund_fields["return_refund_amount"],
            }
        return self._extract_refund_summary_fields(refund_summary)

    def _extract_refund_summary_fields(self, refund_summary: dict[str, Any]) -> dict[str, Any]:
        """
        从退款汇总结构中提取横向输出字段。
        """
        categories = refund_summary.get("categories", {})
        return_refund = categories.get("退货退款", {})
        shipped_only = categories.get("已发货仅退款", {})
        unshipped_only = categories.get("未发货仅退款", {})

        return {
            "return_refund_count": int(return_refund.get("count", 0) or 0),
            "return_refund_amount": round(float(return_refund.get("amount", 0.0) or 0.0), 2),
            "shipped_only_refund_count": int(shipped_only.get("count", 0) or 0),
            "shipped_only_refund_amount": round(float(shipped_only.get("amount", 0.0) or 0.0), 2),
            "unshipped_only_refund_count": int(unshipped_only.get("count", 0) or 0),
            "unshipped_only_refund_amount": round(float(unshipped_only.get("amount", 0.0) or 0.0), 2),
        }

    def _extract_douyin_refund_summary_fields(self, refund_summary: dict[str, Any]) -> dict[str, Any]:
        """
        从抖音罗盘退款分析汇总结构中提取横向输出字段。
        """
        metrics = refund_summary.get("douyin_refund_metrics", {})

        def get_count(key: str) -> int:
            return int(round(float(metrics.get(key, 0) or 0)))

        def get_amount(key: str) -> float:
            return round(float(metrics.get(key, 0.0) or 0.0), 2)

        return {
            "refund_total_order_count": get_count("refund_total_order_count"),
            "refund_total_amount": get_amount("refund_total_amount"),
            "pre_shipment_refund_order_count": get_count("pre_shipment_refund_order_count"),
            "pre_shipment_refund_amount": get_amount("pre_shipment_refund_amount"),
            "unreceived_refund_order_count": get_count("unreceived_refund_order_count"),
            "unreceived_refund_amount": get_amount("unreceived_refund_amount"),
            "received_refund_order_count": get_count("received_refund_order_count"),
            "received_refund_amount": get_amount("received_refund_amount"),
            "return_refund_order_count": get_count("return_refund_order_count"),
            "return_refund_amount": get_amount("return_refund_amount"),
        }

    def _write_refund_summary_table(
        self,
        worksheet: Worksheet,
        refund_summary: dict[str, Any],
    ) -> None:
        """
        在无模板场景下，写入退款汇总横向表格。
        """
        worksheet.title = "退款汇总"
        refund_fields = self._extract_refund_summary_fields(refund_summary)

        headers = [
            "退货退款数量",
            "退货退款金额",
            "已发货仅退款数量",
            "已发货仅退款金额",
            "未发货仅退款数量",
            "未发货仅退款金额",
        ]
        values = [
            refund_fields["return_refund_count"],
            refund_fields["return_refund_amount"],
            refund_fields["shipped_only_refund_count"],
            refund_fields["shipped_only_refund_amount"],
            refund_fields["unshipped_only_refund_count"],
            refund_fields["unshipped_only_refund_amount"],
        ]

        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column_index, value=header)
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=2, column=column_index, value=value)
            if "金额" in headers[column_index - 1]:
                cell.number_format = "0.00"
            if "数量" in headers[column_index - 1]:
                cell.number_format = "0"

        for column_index, header in enumerate(headers, start=1):
            column_letter = worksheet.cell(row=1, column=column_index).column_letter
            worksheet.column_dimensions[column_letter].width = max(len(str(header)) + 2, 14)

    def export_refund_with_business_finance_metrics(
        self,
        refund_summary: dict[str, Any],
        metrics: dict[str, Any],
        output_path: Optional[Path] = None,
        summary_sheet_name: str = "",
        metrics_sheet_name: str = "业务财务汇总",
    ) -> Path:
        """
        输出“退款模板报表 + 业务财务汇总”到同一个工作簿。
        """
        try:
            self.load_template()
            self.prepare_sheet(sheet_name=summary_sheet_name)
            self.write_summary(stats=refund_summary)
        except FileNotFoundError:
            # 模板不存在时，自动降级为内置结构，避免流程中断。
            self.create_workbook()
            if self.worksheet is None:
                raise RuntimeError("工作簿未初始化，无法写入退款汇总。")
            self._write_refund_summary_table(worksheet=self.worksheet, refund_summary=refund_summary)

        if self.workbook is None:
            raise RuntimeError("工作簿未初始化，请先调用 load_template。")

        if metrics_sheet_name in self.workbook.sheetnames:
            metrics_sheet = self.workbook[metrics_sheet_name]
            max_row = metrics_sheet.max_row
            if max_row > 0:
                metrics_sheet.delete_rows(1, max_row)
        else:
            metrics_sheet = self.workbook.create_sheet(title=metrics_sheet_name)

        self._write_business_finance_metrics_sheet(
            worksheet=metrics_sheet,
            metrics=metrics,
            refund_summary=refund_summary,
        )
        report_date = str(metrics.get("report_date", DateConfig.default_report_date_str()))
        shop_name = self._build_platform_prefixed_name(
            shop_name=str(metrics.get("shop_name", "")).strip(),
            platform=metrics.get("platform", ""),
            default_prefix="报表",
        )
        if output_path is None:
            target_dir = Path.home() / "Desktop"
        else:
            path_obj = Path(output_path)
            target_dir = path_obj if path_obj.suffix == "" else path_obj.parent
        target_file = self._build_output_file_path(
            output_path=target_dir,
            report_date=report_date,
            shop_name=shop_name,
            default_prefix="报表",
        )
        self.workbook.save(target_file)
        return target_file
