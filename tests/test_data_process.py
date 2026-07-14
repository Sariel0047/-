"""
`data_process.py` 测试骨架。
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pandas as pd

import qianiu_auto_report.data_process as data_process_module
from qianiu_auto_report.data_process import DataProcessor


def test_placeholder() -> None:
    """
    占位测试。
    """
    pass


def test_validate_columns_prefers_refund_finished_time_column() -> None:
    """
    同时存在申请时间与退款完结时间时，应优先按退款完结时间筛选。
    """
    df = pd.DataFrame(
        {
            "申请时间": ["2026-03-30 08:00:00"],
            "退款完结时间": ["2026-03-31 09:00:00"],
            "货物状态": ["未发货"],
            "退款总额": [12.5],
        }
    )
    processor = DataProcessor()
    processor.validate_columns(df)
    assert processor.date_column == "退款完结时间"


def test_summarize_refund_metrics_by_goods_status_and_finished_date(monkeypatch) -> None:
    """
    按“退款完结时间=昨天”筛选后，统计三类货物状态的数量与退款总额。
    """
    monkeypatch.setattr(
        data_process_module,
        "get_previous_date",
        lambda: date(2026, 3, 31),
    )

    raw_df = pd.DataFrame(
        {
            "退款完结时间": [
                "2026-03-31 08:00:00",
                "2026-03-31 09:00:00",
                "2026-03-31 10:00:00",
                "2026-03-31 11:00:00",
                "2026-03-30 12:00:00",
            ],
            "货物状态": [
                "未收到货",
                "已收到货",
                "未发货",
                "已寄回",
                "未收到货",
            ],
            "退款总额": [
                "¥10.00",
                "20.00",
                "30",
                "40",
                "999",
            ],
            "售后类型": [
                "其他",
                "其他",
                "其他",
                "其他",
                "其他",
            ],
        }
    )

    processor = DataProcessor()
    cleaned_df = processor.clean_data(raw_df)
    processor.validate_columns(cleaned_df)
    transformed_df = processor.transform_data(cleaned_df)
    summary = processor.summarize_data(transformed_df)

    shipped_only_refund = summary["categories"]["已发货仅退款"]
    unshipped_only_refund = summary["categories"]["未发货仅退款"]
    return_refund = summary["categories"]["退货退款"]

    assert shipped_only_refund["count"] == 2
    assert shipped_only_refund["amount"] == 30.0

    assert unshipped_only_refund["count"] == 1
    assert unshipped_only_refund["amount"] == 30.0

    assert return_refund["count"] == 1
    assert return_refund["amount"] == 40.0


def test_process_filters_by_explicit_report_date(tmp_path: Path) -> None:
    """
    指定报表日期时，应按该日期筛选退款完结时间，而不是默认昨天。
    """
    input_file = tmp_path / "refunds.xlsx"
    pd.DataFrame(
        {
            "退款完结时间": [
                "2026-03-29 08:00:00",
                "2026-03-31 09:00:00",
            ],
            "货物状态": ["未发货", "已寄回"],
            "退款总额": ["10.00", "99.00"],
            "售后类型": ["其他", "其他"],
        }
    ).to_excel(input_file, index=False)

    summary = DataProcessor().process(
        input_path=input_file,
        output_path=tmp_path / "processed.xlsx",
        report_date=date(2026, 3, 29),
    )

    assert summary["total_count"] == 1
    assert summary["total_amount"] == 10.0
    assert summary["categories"]["未发货仅退款"]["count"] == 1


def test_supported_table_suffixes_include_wps_et() -> None:
    """
    手动导入和浏览器下载处理都应接受 WPS 表格 .et 文件。
    """
    assert ".et" in DataProcessor.SUPPORTED_TABLE_SUFFIXES


def test_read_table_file_uses_xlrd_for_legacy_excel_and_wps_et(
    tmp_path: Path,
    monkeypatch,
) -> None:
    """
    .xls 和 WPS .et 应走 xlrd 引擎，避免 pandas 默认缺引擎导致读取失败。
    """
    calls: list[tuple[Path, str | None, object]] = []

    def fake_read_excel(path: Path, *args: object, **kwargs: object) -> pd.DataFrame:
        calls.append((Path(path), kwargs.get("engine"), kwargs.get("dtype")))
        return pd.DataFrame({"商品ID": ["P1"]})

    monkeypatch.setattr(pd, "read_excel", fake_read_excel)

    processor = DataProcessor()
    xls_file = tmp_path / "orders.xls"
    et_file = tmp_path / "orders.et"
    xls_file.write_bytes(b"placeholder")
    et_file.write_bytes(b"placeholder")

    processor._read_table_file(xls_file)
    processor._read_table_file(et_file)

    assert calls == [
        (xls_file, "xlrd", str),
        (et_file, "xlrd", str),
    ]


def test_summarize_douyin_refund_analysis_detail_by_stage_columns(tmp_path) -> None:
    """
    抖音罗盘退款分析明细应按“本店数据”阶段列求和。
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "本店数据"
    worksheet.append(
        [
            "日期",
            "退款口径",
            "售卖类型",
            "载体",
            "全部退款阶段-退款人数",
            "全部退款阶段-退款订单数",
            "全部退款阶段-退款金额",
            "全部退款阶段-退款率",
            "发货前退款阶段-退款人数",
            "发货前退款阶段-退款订单数",
            "发货前退款阶段-退款金额",
            "发货前退款阶段-退款率",
            "未收货退款阶段-退款人数",
            "未收货退款阶段-退款订单数",
            "未收货退款阶段-退款金额",
            "未收货退款阶段-退款率",
            "已收货退款阶段-退款人数",
            "已收货退款阶段-退款订单数",
            "已收货退款阶段-退款金额",
            "已收货退款阶段-退款率",
            "已收货退货退款阶段-退款人数",
            "已收货退货退款阶段-退款订单数",
            "已收货退货退款阶段-退款金额",
            "已收货退货退款阶段-退款率",
        ]
    )
    worksheet.append(
        [
            "2026/05/13",
            "订单支付时间",
            "自营",
            "商品卡",
            10,
            11,
            "¥100.50",
            "10%",
            8,
            9,
            "80.25",
            "8%",
            1,
            2,
            "20",
            "2%",
            3,
            4,
            "40",
            "4%",
            5,
            6,
            "60",
            "6%",
        ]
    )
    worksheet.append(
        [
            "2026/05/13",
            "订单支付时间",
            "合作",
            "短视频",
            20,
            21,
            200,
            "20%",
            18,
            19,
            180,
            "18%",
            11,
            12,
            120,
            "12%",
            13,
            14,
            140,
            "14%",
            15,
            16,
            160,
            "16%",
        ]
    )
    input_file = tmp_path / "douyin_refund.xlsx"
    workbook.save(input_file)

    summary = DataProcessor().summarize_douyin_refund_analysis(input_file)
    refund_metrics = summary["douyin_refund_metrics"]

    assert summary["report_date"] == "2026-05-13"
    assert refund_metrics["refund_total_order_count"] == 32
    assert refund_metrics["refund_total_amount"] == 300.5
    assert refund_metrics["pre_shipment_refund_order_count"] == 28
    assert refund_metrics["pre_shipment_refund_amount"] == 260.25
    assert refund_metrics["unreceived_refund_order_count"] == 14
    assert refund_metrics["unreceived_refund_amount"] == 140.0
    assert refund_metrics["received_refund_order_count"] == 18
    assert refund_metrics["received_refund_amount"] == 180.0
    assert refund_metrics["return_refund_order_count"] == 22
    assert refund_metrics["return_refund_amount"] == 220.0


def test_summarize_douyin_refund_analysis_empty_detail_defaults_to_zero(tmp_path) -> None:
    """
    抖音退款分析明细无数据时，缺少阶段列也应按 0 汇总，而不是中断流程。
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "本店数据"
    worksheet.append(["暂无数据"])
    input_file = tmp_path / "empty_douyin_refund.xlsx"
    workbook.save(input_file)

    summary = DataProcessor().summarize_douyin_refund_analysis(input_file)
    refund_metrics = summary["douyin_refund_metrics"]

    assert refund_metrics == {
        "refund_total_order_count": 0,
        "refund_total_amount": 0.0,
        "pre_shipment_refund_order_count": 0,
        "pre_shipment_refund_amount": 0.0,
        "unreceived_refund_order_count": 0,
        "unreceived_refund_amount": 0.0,
        "received_refund_order_count": 0,
        "received_refund_amount": 0.0,
        "return_refund_order_count": 0,
        "return_refund_amount": 0.0,
    }
    assert summary["total_count"] == 0
    assert summary["total_amount"] == 0.0


def test_summarize_douyin_refund_analysis_uses_title_date_when_empty(tmp_path) -> None:
    """
    抖音报表日期应以退款明细文件标题中的日期为准，哪怕明细无数据。
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "本店数据"
    worksheet.append(["暂无数据"])
    input_file = tmp_path / "抖音电商罗盘-退款分析-2026_05_13.xlsx"
    workbook.save(input_file)

    summary = DataProcessor().summarize_douyin_refund_analysis(input_file)

    assert summary["report_date"] == "2026-05-13"
    assert summary["douyin_refund_metrics"]["refund_total_order_count"] == 0


def test_summarize_douyin_after_sale_orders_by_after_sale_type(tmp_path) -> None:
    """
    抖音售后工作台导出的售后单应按“售后类型”和“退商品金额（元）”汇总三类退款。
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(
        [
            "售后单号",
            "商品发货状态",
            "售后类型",
            "退商品金额（元）",
            "售后状态",
            "售后完结时间",
        ]
    )
    worksheet.append(["A001", "1/1已发货", "退货退款", "159", "同意退款，退款成功", "2026-05-16 15:03:52"])
    worksheet.append(["A002", "1/1已发货", "已发货退款", "¥89.50", "同意退款，退款成功", "2026-05-16 18:00:00"])
    worksheet.append(["A003", "未发货", "未发货退款", "100.25", "同意退款，退款成功", "2026-05-16 20:00:00"])
    worksheet.append(["A004", "未发货", "未发货退款", "-", "同意退款，退款成功", "2026-05-16 21:00:00"])
    input_file = tmp_path / "售后单-2026-05-18 00_34_04.xlsx"
    workbook.save(input_file)

    summary = DataProcessor().summarize_douyin_after_sale_orders(input_file)
    refund_metrics = summary["douyin_refund_metrics"]

    assert summary["report_date"] == "2026-05-16"
    assert refund_metrics["refund_total_order_count"] == 4
    assert refund_metrics["refund_total_amount"] == 348.75
    assert refund_metrics["pre_shipment_refund_order_count"] == 2
    assert refund_metrics["pre_shipment_refund_amount"] == 100.25
    assert refund_metrics["received_refund_order_count"] == 1
    assert refund_metrics["received_refund_amount"] == 89.5
    assert refund_metrics["return_refund_order_count"] == 1
    assert refund_metrics["return_refund_amount"] == 159.0

    assert summary["categories"]["未发货仅退款"]["count"] == 2
    assert summary["categories"]["未发货仅退款"]["amount"] == 100.25
    assert summary["categories"]["已发货仅退款"]["count"] == 1
    assert summary["categories"]["已发货仅退款"]["amount"] == 89.5
    assert summary["categories"]["退货退款"]["count"] == 1
    assert summary["categories"]["退货退款"]["amount"] == 159.0


def test_summarize_douyin_after_sale_orders_empty_uses_previous_date(
    tmp_path,
    monkeypatch,
) -> None:
    """
    售后单文件名是下载时间；空表时统计日期应按“昨日”，不能误用文件名日期。
    """
    monkeypatch.setattr(
        data_process_module,
        "get_previous_date",
        lambda: date(2026, 5, 17),
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["售后类型", "退商品金额（元）", "售后完结时间"])
    input_file = tmp_path / "售后单-2026-05-18 00_34_04.xlsx"
    workbook.save(input_file)

    summary = DataProcessor().summarize_douyin_after_sale_orders(input_file)

    assert summary["report_date"] == "2026-05-17"
    assert summary["douyin_refund_metrics"]["refund_total_order_count"] == 0


def test_summarize_tmall_sold_orders_by_product_id_and_refund_type(tmp_path: Path) -> None:
    """
    已卖出宝贝导出的宝贝销售明细应按商品 ID 汇总明细行数、仅退款与退货退款。
    """
    input_file = tmp_path / "tmall_orders.xlsx"
    pd.DataFrame(
        {
            "主订单编号": ["A001", "A001", "A002", "A003", "B001", "B002"],
            "购买数量": [2, 3, 4, 5, 1, 2],
            "订单状态": ["交易成功", "交易成功", "交易关闭", "交易失败", "交易成功", "交易失败"],
            "买家实付金额": [60, 40, 30, 20, 80, 10],
            "商品ID": ["P1", "P1", "P1", "P1", "P2", "P2"],
            "物流单号": ["Y001", "Y001", "", "Y003", "Y101", ""],
        }
    ).to_excel(input_file, index=False)

    summary_df = DataProcessor().summarize_tmall_sold_orders(input_file)

    assert summary_df.columns.tolist() == [
        "商品id",
        "订单笔数",
        "订单金额",
        "仅退款笔数",
        "仅退款金额",
        "实际发出笔数",
        "实际发出金额",
        "退货退款笔数",
        "退货退款金额",
        "实际成交笔数",
    ]

    rows = {row["商品id"]: row for row in summary_df.to_dict("records")}
    assert rows["P1"] == {
        "商品id": "P1",
        "订单笔数": 4,
        "订单金额": 150.0,
        "仅退款笔数": 1,
        "仅退款金额": 30.0,
        "实际发出笔数": 3,
        "实际发出金额": 120.0,
        "退货退款笔数": 1,
        "退货退款金额": 20.0,
        "实际成交笔数": 2,
    }
    assert rows["P2"] == {
        "商品id": "P2",
        "订单笔数": 2,
        "订单金额": 90.0,
        "仅退款笔数": 1,
        "仅退款金额": 10.0,
        "实际发出笔数": 1,
        "实际发出金额": 80.0,
        "退货退款笔数": 0,
        "退货退款金额": 0.0,
        "实际成交笔数": 1,
    }


def test_summarize_tmall_sold_orders_filters_to_requested_product_ids(tmp_path: Path) -> None:
    """
    导出的明细可能带出同订单其它商品，汇总时只保留用户查询的商品 ID。
    """
    input_file = tmp_path / "tmall_orders.xlsx"
    pd.DataFrame(
        {
            "主订单编号": ["A001", "A001", "A002", "A003"],
            "购买数量": [1, 1, 1, 1],
            "订单状态": ["交易成功", "交易成功", "交易关闭", "交易成功"],
            "买家实付金额": [60, 40, 30, 20],
            "商品ID": ["P1", "EXTRA", "P2", "P3"],
            "物流单号": ["Y001", "Y001", "", "Y003"],
        }
    ).to_excel(input_file, index=False)

    summary_df = DataProcessor().summarize_tmall_sold_orders(
        input_file,
        product_ids="P2,P1",
    )

    assert summary_df["商品id"].tolist() == ["P2", "P1"]
    assert "EXTRA" not in summary_df["商品id"].tolist()
    assert "P3" not in summary_df["商品id"].tolist()


def test_save_tmall_sold_order_summary_writes_excel(tmp_path: Path) -> None:
    """
    天猫订单汇总应能保存为新的 Excel 文件。
    """
    input_file = tmp_path / "tmall_orders.xlsx"
    output_file = tmp_path / "tmall_summary.xlsx"
    pd.DataFrame(
        {
            "主订单编号": ["A001"],
            "购买数量": [1],
            "订单状态": ["交易成功"],
            "买家实付金额": [119],
            "商品ID": ["906669497660"],
            "物流单号": ["79015758079719"],
        }
    ).to_excel(input_file, index=False)

    result = DataProcessor().save_tmall_sold_order_summary(input_file, output_file)

    assert result == output_file
    workbook = load_workbook(output_file)
    worksheet = workbook.active
    assert [worksheet.cell(row=1, column=index).value for index in range(1, 5)] == [
        "商品id",
        None,
        None,
        "订单笔数",
    ]
    assert worksheet.cell(row=2, column=1).value == "906669497660"
    assert worksheet.cell(row=2, column=4).value == 1
    assert worksheet.cell(row=2, column=5).value == 119

    amount_headers = {"订单金额", "仅退款金额", "实际发出金额", "退货退款金额"}
    amount_columns = [
        cell.column
        for cell in worksheet[1]
        if cell.value in amount_headers
    ]
    assert amount_columns
    for column_index in amount_columns:
        assert worksheet.cell(row=2, column=column_index).number_format == "0.00"


def test_summarize_douyin_order_details_by_product_id(tmp_path: Path) -> None:
    """
    抖音订单明细离线表应按商品 ID 汇总整份文件，并按订单笔数倒序。
    """
    input_file = tmp_path / "douyin_orders.xlsx"
    pd.DataFrame(
        {
            "主订单编号": ["A001", "A002", "A003", "A004", "B001"],
            "购买数量": [2, 1, 3, 4, 5],
            "订单状态": ["交易成功", "交易关闭", "交易关闭", "交易成功", "交易关闭"],
            "买家实付金额": [200, 80, 150, 120, 300],
            "售后状态": ["-", "退款成功", "退款成功", "-", "退款成功"],
            "退款金额": ["无退款申请", 1, 1, "无退款申请", 1],
            "商品ID": ["P1", "P1", "P1", "P1", "P2"],
            "发货时间": [
                "2026-06-16 10:00:00",
                "",
                "2026-06-16 11:00:00",
                "2026-06-16 12:00:00",
                "",
            ],
        }
    ).to_excel(input_file, index=False)

    summary_df = DataProcessor().summarize_douyin_order_details(input_file)

    assert summary_df["商品id"].tolist() == ["P1", "P2"]
    rows = {row["商品id"]: row for row in summary_df.to_dict("records")}
    assert rows["P1"] == {
        "商品id": "P1",
        "订单笔数": 10,
        "订单金额": 550.0,
        "仅退款笔数": 1,
        "仅退款金额": 80.0,
        "实际发出笔数": 9,
        "实际发出金额": 470.0,
        "退货退款笔数": 3,
        "退货退款金额": 150.0,
        "实际成交笔数": 6,
    }
    assert rows["P2"]["订单笔数"] == 5
    assert rows["P2"]["仅退款金额"] == 300.0


def test_summarize_douyin_order_details_accepts_standard_csv(tmp_path: Path) -> None:
    """
    抖音标准报表 CSV 应按订单状态、订单应付金额和发货时间汇总。
    """
    input_file = tmp_path / "douyin_orders.csv"
    pd.DataFrame(
        {
            "主订单编号": ["A001", "A002", "A003"],
            "商品数量": [1, 2, 3],
            "订单应付金额": [100, 200, 300],
            "商品ID": ["P1", "P1", "P2"],
            "订单状态": ["已关闭", "已关闭", "已完成"],
            "售后状态": ["退款成功", "退款成功", "-"],
            "发货时间": ["-", "2026-06-16 12:00:00", "2026-06-17 12:00:00"],
        }
    ).to_csv(input_file, index=False, encoding="utf-8-sig")

    summary_df = DataProcessor().summarize_douyin_order_details(input_file)

    rows = {row["商品id"]: row for row in summary_df.to_dict("records")}
    assert rows["P2"]["订单笔数"] == 3
    assert rows["P1"]["订单笔数"] == 3
    assert rows["P1"]["仅退款笔数"] == 1
    assert rows["P1"]["仅退款金额"] == 100.0
    assert rows["P1"]["退货退款笔数"] == 2
    assert rows["P1"]["退货退款金额"] == 200.0


def test_summarize_douyin_order_details_uses_after_sale_status_not_order_status(
    tmp_path: Path,
) -> None:
    """
    抖音离线表应按售后状态判断退款，订单状态不能单独把已关闭订单算成退款。
    """
    input_file = tmp_path / "douyin_orders.csv"
    pd.DataFrame(
        {
            "商品数量": [1, 2, 3],
            "订单应付金额": [10.55, 20.25, 30.75],
            "商品ID": ["P1", "P1", "P1"],
            "订单状态": ["已完成", "已关闭", "已关闭"],
            "售后状态": ["退款成功", "-", "同意退款，退款成功"],
            "退款金额": [999, 999, 999],
            "发货时间": ["", "", "2026-06-16 12:00:00"],
        }
    ).to_csv(input_file, index=False, encoding="utf-8-sig")

    summary_df = DataProcessor().summarize_douyin_order_details(input_file)

    row = summary_df.to_dict("records")[0]
    assert row["订单笔数"] == 6
    assert row["订单金额"] == 61.55
    assert row["仅退款笔数"] == 1
    assert row["仅退款金额"] == 10.55
    assert row["退货退款笔数"] == 3
    assert row["退货退款金额"] == 30.75


def test_save_douyin_order_detail_summary_sets_amount_column_width(tmp_path: Path) -> None:
    """
    抖音汇总表金额列应有足够列宽，避免 WPS/Excel 显示为 ######。
    """
    input_file = tmp_path / "douyin_orders.csv"
    output_file = tmp_path / "douyin_summary.xlsx"
    pd.DataFrame(
        {
            "主订单编号": ["A001"],
            "商品数量": [10168],
            "订单应付金额": [1288888.88],
            "商品ID": ["3764109164402573517"],
            "订单状态": ["已完成"],
            "售后状态": ["-"],
            "发货时间": ["2026-06-16 12:00:00"],
        }
    ).to_csv(input_file, index=False, encoding="utf-8-sig")

    DataProcessor().save_douyin_order_detail_summary(input_file, output_file)

    worksheet = load_workbook(output_file).active
    amount_column = next(cell.column_letter for cell in worksheet[1] if cell.value == "订单金额")
    assert worksheet.column_dimensions[amount_column].width >= 14
    assert worksheet[f"{amount_column}2"].number_format == "0.00"
