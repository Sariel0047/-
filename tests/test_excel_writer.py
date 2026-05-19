"""
`excel_writer.py` 测试骨架。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from qianiu_auto_report.excel_writer import ExcelWriter


def test_placeholder() -> None:
    """
    占位测试。
    """
    pass


def test_export_refund_with_business_metrics_fallback_without_template(tmp_path: Path) -> None:
    """
    模板不存在时，应自动使用内置结构输出“退款汇总 + 业务财务汇总”。
    """
    writer = ExcelWriter(template_path=tmp_path / "template_missing.xlsx")

    refund_summary = {
        "total_count": 4,
        "total_amount": 100.0,
        "categories": {
            "已发货仅退款": {"count": 2, "amount": 30.0, "sub_categories": {}},
            "未发货仅退款": {"count": 1, "amount": 30.0, "sub_categories": {}},
            "退货退款": {"count": 1, "amount": 40.0, "sub_categories": {}},
        },
    }
    metrics = {
        "report_date": "2026-03-31",
        "shop_name": "vullvan瑜妍旗舰店",
        "payment_buyer_count": 10,
        "payment_amount": 88.8,
        "payment_sub_order_count": 12,
        "trade_compensation": 11.9,
        "cross_border_value_added_fee": 20.8,
        "promotion_fee": 1122.33,
    }

    output_file = writer.export_refund_with_business_finance_metrics(
        refund_summary=refund_summary,
        metrics=metrics,
        output_path=tmp_path,
    )

    assert output_file.exists()
    assert output_file.name == "vullvan瑜妍旗舰店_2026-03-31.xlsx"

    workbook = load_workbook(output_file)
    assert "退款汇总" in workbook.sheetnames
    assert "业务财务汇总" in workbook.sheetnames

    refund_sheet = workbook["退款汇总"]
    assert refund_sheet["A1"].value == "退货退款数量"
    assert refund_sheet["B1"].value == "退货退款金额"
    assert refund_sheet["C1"].value == "已发货仅退款数量"
    assert refund_sheet["D1"].value == "已发货仅退款金额"
    assert refund_sheet["E1"].value == "未发货仅退款数量"
    assert refund_sheet["F1"].value == "未发货仅退款金额"
    assert refund_sheet["A2"].value == 1
    assert float(refund_sheet["B2"].value) == 40.0
    assert refund_sheet["C2"].value == 2
    assert float(refund_sheet["D2"].value) == 30.0
    assert refund_sheet["E2"].value == 1
    assert float(refund_sheet["F2"].value) == 30.0

    finance_sheet = workbook["业务财务汇总"]
    headers = [finance_sheet.cell(row=1, column=column).value for column in range(1, 14)]
    values = [finance_sheet.cell(row=2, column=column).value for column in range(1, 14)]
    assert headers == [
        "统计日期",
        "支付买家数",
        "支付子订单数",
        "支付金额",
        "未发货仅退款",
        "未发货仅退款金额",
        "已发货仅退款数量",
        "已发货仅退款金额",
        "退货退款数量",
        "退货退款金额",
        "交易赔付",
        "淘宝天猫跨境服务增值费",
        "推广费用",
    ]
    assert values == [
        "2026-03-31",
        10,
        12,
        88.8,
        1,
        30.0,
        2,
        30.0,
        1,
        40.0,
        11.9,
        20.8,
        1122.33,
    ]


def test_export_business_finance_metrics_uses_shop_name_filename(tmp_path: Path) -> None:
    """
    业务财务汇总输出应优先使用“平台_店铺名_日期.xlsx”命名。
    """
    writer = ExcelWriter()
    metrics = {
        "report_date": "2026-04-01",
        "shop_name": "vullvan/瑜妍旗舰店",
        "platform": "taobao",
        "payment_buyer_count": 1,
        "payment_amount": 1.0,
        "payment_sub_order_count": 1,
        "trade_compensation": 0.0,
        "cross_border_value_added_fee": 0.0,
        "promotion_fee": 0.0,
    }

    output_file = writer.export_business_finance_metrics(metrics=metrics, output_path=tmp_path)

    assert output_file.exists()
    # "/" 会被清洗为 "_"
    assert output_file.name == "淘宝_vullvan_瑜妍旗舰店_2026-04-01.xlsx"


def test_export_business_finance_metrics_writes_douyin_refund_metrics(tmp_path: Path) -> None:
    """
    抖音流程的业务财务汇总应写入罗盘退款分析明细汇总字段。
    """
    writer = ExcelWriter()
    metrics = {
        "report_date": "2026-05-13",
        "shop_name": "高品质裙裤",
        "platform": "douyin",
        "payment_buyer_count": 784,
        "payment_amount": 108936.0,
        "payment_sub_order_count": 784,
        "trade_compensation": 0.0,
        "cross_border_value_added_fee": 0.0,
        "promotion_fee": 1745.05,
        "refund_summary": {
            "douyin_refund_metrics": {
                "refund_total_order_count": 185,
                "refund_total_amount": 25809.77,
                "pre_shipment_refund_order_count": 11,
                "pre_shipment_refund_amount": 100.1,
                "unreceived_refund_order_count": 3,
                "unreceived_refund_amount": 30.3,
                "received_refund_order_count": 5,
                "received_refund_amount": 50.5,
                "return_refund_order_count": 7,
                "return_refund_amount": 70.7,
            }
        },
    }

    output_file = writer.export_business_finance_metrics(metrics=metrics, output_path=tmp_path)
    assert output_file.name == "抖音_高品质裙裤_2026-05-13.xlsx"
    workbook = load_workbook(output_file)
    sheet = workbook["业务财务汇总"]
    headers = [sheet.cell(row=1, column=column).value for column in range(1, 14)]
    values = [sheet.cell(row=2, column=column).value for column in range(1, 14)]

    assert headers == [
        "统计日期",
        "支付买家数",
        "支付子订单数",
        "支付金额",
        "未发货仅退款",
        "未发货仅退款金额",
        "已发货仅退款数量",
        "已发货仅退款金额",
        "退货退款数量",
        "退货退款金额",
        "交易赔付",
        "淘宝天猫跨境服务增值费",
        "推广费用",
    ]
    assert values == [
        "2026-05-13",
        784,
        784,
        108936.0,
        11,
        100.1,
        8,
        80.8,
        7,
        70.7,
        0.0,
        0.0,
        1745.05,
    ]
